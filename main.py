from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QFrame,
    QPushButton, QLabel, QLineEdit, QDateEdit, QTimeEdit,
    QComboBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QApplication, QMenuBar, QMenu, QFileDialog, QMessageBox,
    QDialog
)
from PySide6.QtCore import Qt, QTime, QDate
from PySide6.QtGui import QFont, QColor
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtWidgets import QFormLayout, QTabWidget, QSpinBox
from PySide6.QtWidgets import QTableWidget
from PySide6.QtGui import QBrush, QColor
import sys
import json
import calendar
import sqlite3
import os
from database_manager import DatabaseManager
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

class MenuBar(QMenuBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.createFileMenu()
        self.createDatabaseMenu()

    def createFileMenu(self):
        fileMenu = self.addMenu("Fájl")
        fileMenu.addAction("Excel megnyitása").triggered.connect(lambda: self.parent().openExcel())
        fileMenu.addAction("Mentés").triggered.connect(lambda: self.parent().saveWorkHoursToExcel())
        fileMenu.addAction("Nyomtatás").triggered.connect(lambda: self.parent().printData())
        fileMenu.addAction("Kilépés").triggered.connect(self.parent().close)

    def createDatabaseMenu(self):
        dbMenu = self.addMenu("Adatbázis")
        dbMenu.addAction("Törzsadatok kezelése").triggered.connect(lambda: self.parent().openDatabaseManager())

class FuvarAdminApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initDatabase()
        self.initUI()

    def createInputGroup(self, label_text, widget):
        layout = QHBoxLayout()
        label = QLabel(label_text)
        label.setStyleSheet(self.styles['label'])
        layout.addWidget(label)
        widget.setStyleSheet(self.styles['input'])
        layout.addWidget(widget)
        layout.addStretch()
        return layout

    def initUI(self):
        self.setWindowTitle("Fuvar Adminisztráció")  
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout()
        self.driver_combo = QComboBox()
        self.vehicle_combo = QComboBox()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        self.setMenuBar(MenuBar(self))
        self.setupStyles()
        self.work_table = QTableWidget() 
        self.delivery_table = QTableWidget()
        self.work_table.setColumnCount(0)  
        self.work_table.setRowCount(0)
        self.setupTopFrame(main_layout)
        self.setupBottomFrame(main_layout)
        main_widget.setLayout(main_layout)
        self.loadSavedWorkHours()  # Add this line
        self.showMaximized()

    def loadDrivers(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT name FROM drivers ORDER BY name")
            drivers = cursor.fetchall()
            self.driver_combo.clear()
            self.driver_combo.addItems([driver[0] for driver in drivers])
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba a sofőrök betöltésekor: {str(e)}")

    def loadVehicles(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT plate_number FROM vehicles ORDER BY plate_number")
            vehicles = cursor.fetchall()
            self.vehicle_combo.clear()
            self.vehicle_combo.addItems([vehicle[0] for vehicle in vehicles])
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba a járművek betöltésekor: {str(e)}")

    def initDatabase(self):
        self.conn = sqlite3.connect('fuvarok.db')
        cursor = self.conn.cursor()
    
        # Gyárak alaptábla
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS factories (
                id INTEGER PRIMARY KEY,
                nev TEXT
            )
        ''')
    
        # Gyár-övezet árak tábla
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS factory_zone_prices (
                id INTEGER PRIMARY KEY,
                factory_id INTEGER,
                zone_name TEXT,
                price INTEGER,
                FOREIGN KEY (factory_id) REFERENCES factories(id) ON DELETE CASCADE
            )
        ''')
    
        # Gyár állásidő díjak tábla
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS factory_waiting_fees (
                id INTEGER PRIMARY KEY,
                factory_id INTEGER,
                price_per_15_min INTEGER,
                FOREIGN KEY (factory_id) REFERENCES factories(id) ON DELETE CASCADE
            )
        ''')

        # Szabadság keret tábla
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS vacation_allowance (
                id INTEGER PRIMARY KEY,
                year INTEGER,
                total_days INTEGER,
                used_days INTEGER DEFAULT 0
            )
        ''')
    
        self.conn.commit()

    def updateVacationDisplay(self):
        cursor = self.conn.cursor()
        current_year = QDate.currentDate().year()
        
        cursor.execute("""
            SELECT total_days, used_days 
            FROM vacation_allowance 
            WHERE year = ?
        """, (current_year,))
        
        result = cursor.fetchone()
        if result:
            total_days, used_days = result
            self.vacation_label.setText(f"Szabadság: {used_days}/{total_days}")
        else:
            cursor.execute("""
                INSERT INTO vacation_allowance (year, total_days, used_days)
                VALUES (?, 29, 0)
            """, (current_year,))
            self.conn.commit()
            self.vacation_label.setText("Szabadság: 0/29")

    def setupStyles(self):
        self.styles = {
            'main_frame': """
                QFrame {
                    background-color: #2d2d2d;
                    border: 3px solid #3e3e3e;
                    border-radius: 15px;
                    margin: 5px;
                }
            """,
            'sub_frame': """
                QFrame {
                    background-color: #d9d9d9;
                    border: 2px solid #4d4d4d;
                    border-radius: 15px;
                    margin: 5px;
                }
            """,
            'table_frame': """
                QFrame {
                    background-color: #d9d9d9;
                    border: 3px solid #ff2800;
                    border-radius: 15px;
                    padding: 10px;
                }
            """,
            'label': """
                QLabel {
                    color: black;
                    font-size: 14px;
                    font-weight: bold;
                    padding: 5px;
                }
            """,
            'input': """
                QLineEdit, QDateEdit, QTimeEdit, QComboBox {
                    background-color: white; 
                    padding: 5px;
                    border: 2px solid #a0a0a0;
                    border-radius: 5px;
                    color: black;
                    min-width: 150px;
                    max-width: 150px;
                    min-height: 30px;
                    max-height: 30px;
                    font-size: 14px;
                }
            """,
            'button': """
                QPushButton {
                    background-color: #4a90e2;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    min-width: 120px;
                }
                QPushButton:hover {
                    background-color: #357abd;
                }
            """
        }

    def setupTableStyles(self):
        table_style = """
            QTableWidget {
                color: black;
            }
            QTableWidget::item {
                color: black;
            }
            QHeaderView::section {
                color: black;
            }
        """
        # Stílus alkalmazása a táblázatokra
        self.work_table.setStyleSheet(table_style)
        self.delivery_table.setStyleSheet(table_style)

    
        for table in [self.work_table, self.delivery_table]:
            table.setStyleSheet(table_style)
        
            # Rács beállítások
            table.setShowGrid(True)
            table.setGridStyle(Qt.SolidLine)
        
            # Fejléc beállítások
            header = table.horizontalHeader()
            header.setVisible(True)
            header.setHighlightSections(True)
            header.setStretchLastSection(True)
            header.setMinimumHeight(40)
        
            # Sorok és oszlopok méretezése
            table.verticalHeader().setDefaultSectionSize(40)
            for i in range(table.columnCount()):
                table.setColumnWidth(i, 150)

    def setupTopFrame(self, main_layout):
        top_frame = QFrame()
        top_frame.setStyleSheet(self.styles['main_frame'])
        top_layout = QHBoxLayout()
    
        # Left Panel - Driver and Vehicle
        left_panel = QFrame()
        left_panel.setStyleSheet(self.styles['sub_frame'])
        left_layout = QVBoxLayout()
    
        # Driver and vehicle selection
        self.driver_combo = QComboBox()
        self.vehicle_combo = QComboBox()
        self.loadDrivers()
        self.loadVehicles()
    
        left_layout.addLayout(self.createInputGroup("Sofőr:", self.driver_combo))
        left_layout.addLayout(self.createInputGroup("Rendszám:", self.vehicle_combo))
        left_layout.addStretch()
    
        left_panel.setLayout(left_layout)
    
        # Middle Panel - Work Hours
        middle_panel = QFrame()
        middle_panel.setStyleSheet(self.styles['sub_frame'])
        middle_layout = QVBoxLayout()
    
        # Date and time inputs
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
    
        self.start_time = QTimeEdit()
        self.end_time = QTimeEdit()
    
        middle_layout.addLayout(self.createInputGroup("Dátum:", self.date_edit))
        middle_layout.addLayout(self.createInputGroup("Kezdés:", self.start_time))
        middle_layout.addLayout(self.createInputGroup("Végzés:", self.end_time))
    
        # Work type and vacation
        self.type_combo = QComboBox()
        self.type_combo.addItems(["Sima munkanap", "Műhely nap", "Szabadság", "Betegszabadság (TP)"])
        middle_layout.addLayout(self.createInputGroup("Munka típusa:", self.type_combo))
    
        self.vacation_label = QLabel("Szabadság: 0/0")
        self.vacation_label.setStyleSheet(self.styles['label'])
        middle_layout.addWidget(self.vacation_label)
    
        middle_panel.setLayout(middle_layout)
    
        # Right Panel - Delivery Data
        right_panel = QFrame()
        right_panel.setStyleSheet(self.styles['sub_frame'])
        right_layout = QVBoxLayout()

        # Munkaórák mentése gomb (bal oldali táblázathoz)
        workhours_button_layout = QHBoxLayout()
        save_workhours_btn = QPushButton("Munkaórák Mentése")
        save_workhours_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF2800;  /* Ferrari piros */
                color: white;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #CC2000;
            }
        """)
        save_workhours_btn.clicked.connect(lambda: self.saveWorkHoursAndExport())
        workhours_button_layout.addWidget(save_workhours_btn)
        workhours_button_layout.addStretch()
        middle_layout.addLayout(workhours_button_layout)

        # Fuvar adatok mentése gomb (jobb oldali táblázathoz)
        delivery_button_layout = QHBoxLayout()
        save_delivery_btn = QPushButton("Fuvar Adatok Mentése")
        save_delivery_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF2800;  /* Ferrari piros */
                color: white;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #CC2000;
            }
        """)
        save_delivery_btn.clicked.connect(lambda: self.saveDeliveryAndExport())
        delivery_button_layout.addWidget(save_delivery_btn)
        delivery_button_layout.addStretch()
        right_layout.addLayout(delivery_button_layout)
    
        # Zone selection and factory
        self.km_combo = QComboBox()
        self.km_combo.addItems([f"Övezet {i}-{i+5}" for i in range(0, 50, 5)])
    
        self.factory_combo = QComboBox()
        self.loadFactories()
    
        right_layout.addLayout(self.createInputGroup("Kilométer sáv:", self.km_combo))
        right_layout.addLayout(self.createInputGroup("Gyár:", self.factory_combo))
    
        # Address and delivery number
        self.address_input = QLineEdit()
        self.delivery_input = QLineEdit()
    
        right_layout.addLayout(self.createInputGroup("Cím:", self.address_input))
        right_layout.addLayout(self.createInputGroup("Szállítószám:", self.delivery_input))
    
        # M3 input
        m3_layout = QHBoxLayout()
        m3_label = QLabel("M3:")
        m3_label.setStyleSheet(self.styles['label'])
        self.m3_input = QLineEdit()
        self.m3_sum = QLabel("(0)")
    
        m3_layout.addWidget(m3_label)
        m3_layout.addWidget(self.m3_input)
        m3_layout.addWidget(self.m3_sum)
        m3_layout.addStretch()
    
        right_layout.addLayout(m3_layout)
        right_panel.setLayout(right_layout)
    
        # Add all panels to top layout
        top_layout.addWidget(left_panel)
        top_layout.addWidget(middle_panel)
        top_layout.addWidget(right_panel)
    
        top_frame.setLayout(top_layout)
        main_layout.addWidget(top_frame)

    def loadFactories(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT nev FROM factories")
        factories = cursor.fetchall()
        self.factory_combo.addItems([factory[0] for factory in factories])

    def setupBottomFrame(self, main_layout):
        bottom_frame = QFrame()
        bottom_frame.setStyleSheet("""
            QFrame {
                background-color: #2d2d2d;
                border: 3px solid #3e3e3e;
                border-radius: 15px;
                margin: 5px;
            }
        """)
        bottom_layout = QHBoxLayout()
    
        # Munkaórák táblázat
        work_frame = QFrame()
        work_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 3px solid #ff2800;
                border-radius: 15px;
                padding: 10px;
            }
            QTableWidget {
                background-color: white;
                color: black;
                gridline-color: #808080;
                border: 1px solid #808080;
                border-radius: 0px;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 5px;
                border-bottom: 1px solid #808080;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                color: black;
                padding: 5px;
                border: 1px solid #808080;
                font-weight: bold;
                font-size: 12px;
            }
            QHeaderView::section:selected {
                background-color: #e0e0e0;
            }
        """)
        work_layout = QVBoxLayout()
    
        self.work_table = QTableWidget()
        work_headers = [
            "Dátum", "Nap", "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", 
            "Ledolgozott\nÓrák", "Műhely\nKezdés", "Műhely\nVégzés", 
            "Ledolgozott\nÓrák", "Szabadság", "Betegszabadság\n(TP)"
        ]
    
        self.work_table.setColumnCount(len(work_headers))
        self.work_table.setHorizontalHeaderLabels(work_headers)
    
        # Fejléc beállítások
        header = self.work_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Fixed)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setMinimumHeight(80)
        self.work_table.verticalHeader().setVisible(False)
    
        # Oszlopszélességek beállítása
        column_widths = [100, 80, 100, 100, 80, 100, 100, 80, 80, 100]
        for i, width in enumerate(column_widths):
            self.work_table.setColumnWidth(i, width)
    
        # Táblázat méretezése
        self.work_table.setMinimumHeight(300)
    
        work_layout.addWidget(self.work_table)
        work_frame.setLayout(work_layout)
    
        # Fuvar táblázat
        delivery_frame = QFrame()
        delivery_frame.setStyleSheet(work_frame.styleSheet())  # Ugyanaz a stílus
        delivery_layout = QVBoxLayout()
    
        self.delivery_table = QTableWidget()
        delivery_headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)]
    
        self.delivery_table.setColumnCount(len(delivery_headers))
        self.delivery_table.setHorizontalHeaderLabels(delivery_headers)
    
        # Fejléc beállítások
        delivery_header = self.delivery_table.horizontalHeader()
        delivery_header.setSectionResizeMode(QHeaderView.Fixed)
        delivery_header.setDefaultAlignment(Qt.AlignCenter)
        delivery_header.setMinimumHeight(80)
        self.delivery_table.verticalHeader().setVisible(False)
    
        # Oszlopszélességek beállítása
        delivery_column_width = 100
        for i in range(self.delivery_table.columnCount()):
            self.delivery_table.setColumnWidth(i, delivery_column_width)
    
        # Táblázat méretezése
        self.delivery_table.setMinimumHeight(300)
    
        delivery_layout.addWidget(self.delivery_table)
        delivery_frame.setLayout(delivery_layout)
    
        # Táblázatok hozzáadása az elrendezéshez
        bottom_layout.addWidget(work_frame)
        bottom_layout.addWidget(delivery_frame)
    
        bottom_frame.setLayout(bottom_layout)
        main_layout.addWidget(bottom_frame)
    
        self.setupTableRows()

    def setupTableRows(self):
        current_date = QDate.currentDate()
        first_day = QDate(current_date.year(), current_date.month(), 1)
        days_in_month = first_day.daysInMonth()

        # Magyar napnevek
        day_names = ['Hétfő', 'Kedd', 'Szerda', 'Csütörtök', 'Péntek', 'Szombat', 'Vasárnap']

        # Munkaóra táblázat sorok
        self.work_table.setRowCount(days_in_month)
        for i in range(days_in_month):
            current_day = first_day.addDays(i)
        
            # Dátum oszlop
            date_item = QTableWidgetItem(current_day.toString('yyyy-MM-dd'))
            date_item.setTextAlignment(Qt.AlignCenter)
            self.work_table.setItem(i, 0, date_item)
        
            # Nap neve oszlop
            day_item = QTableWidgetItem(day_names[current_day.dayOfWeek() - 1])
            day_item.setTextAlignment(Qt.AlignCenter)
            self.work_table.setItem(i, 1, day_item)
        
            # Többi oszlop inicializálása üres cellákkal
            for col in range(2, self.work_table.columnCount()):
                empty_item = QTableWidgetItem("")
                empty_item.setTextAlignment(Qt.AlignCenter)
                self.work_table.setItem(i, col, empty_item)
        
            # Hétvége kiemelése
            if current_day.dayOfWeek() in [6, 7]:  # Szombat és vasárnap
                for col in range(self.work_table.columnCount()):
                    cell = self.work_table.item(i, col)
                    if cell:
                        cell.setBackground(QBrush(QColor("#ffff00")))

    def loadSavedWorkHours(self):
        try:
            if os.path.exists('work_hours.json'):
                with open('work_hours.json', 'r', encoding='utf-8') as f:
                    for line in f:
                        try:
                            data = json.loads(line)
                            date = data.get('date')
                        
                            for row in range(self.work_table.rowCount()):
                                if self.work_table.item(row, 0).text() == date:
                                    # Start time
                                    start_item = QTableWidgetItem(data.get('start_time', ''))
                                    start_item.setTextAlignment(Qt.AlignCenter)
                                    self.work_table.setItem(row, 2, start_item)
                                
                                    # End time
                                    end_item = QTableWidgetItem(data.get('end_time', ''))
                                    end_item.setTextAlignment(Qt.AlignCenter)
                                    self.work_table.setItem(row, 3, end_item)
                                
                                    # Calculate hours
                                    start = QTime.fromString(data.get('start_time', ''), 'HH:mm')
                                    end = QTime.fromString(data.get('end_time', ''), 'HH:mm')
                                    if start.isValid() and end.isValid():
                                        minutes = start.secsTo(end) / 60
                                        hours = minutes / 60
                                        hours_item = QTableWidgetItem(f"{hours:.2f}")
                                        hours_item.setTextAlignment(Qt.AlignCenter)
                                        self.work_table.setItem(row, 4, hours_item)
                                
                                    # Work type
                                    type_item = QTableWidgetItem(data.get('type', ''))
                                    type_item.setTextAlignment(Qt.AlignCenter)
                                    self.work_table.setItem(row, 5, type_item)
                        except json.JSONDecodeError:
                            continue
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Hiba történt az adatok betöltése során: {str(e)}")

    def setupButtons(self, main_layout):
        button_frame = QFrame()
        button_layout = QHBoxLayout()
    
        buttons = [
            ("Munkaórák Mentése", self.saveWorkHours),
            ("Fuvar Adatok Mentése", self.saveDeliveryData),
            ("Munkaórák Excel Export", self.saveWorkHoursToExcel),
            ("Fuvar Excel Export", self.saveDeliveryToExcel),
            ("Kilépés", self.close)
        ]
    
        for text, callback in buttons:
            btn = QPushButton(text)
            btn.clicked.connect(callback)
            btn.setStyleSheet(self.styles['button'])
            button_layout.addWidget(btn)
    
        button_frame.setLayout(button_layout)
        main_layout.addWidget(button_frame)

    def loadFactories(self):
        self.factory_combo.clear()
        cursor = self.conn.cursor()
        cursor.execute("SELECT nev FROM factories")
        factories = cursor.fetchall()
        self.factory_combo.addItems([factory[0] for factory in factories])

    def handleM3Input(self):
        text = self.m3_input.text().strip()
        try:
            # Vesszőt pontra cserélünk
            text = text.replace(',', '.')
        
            if not text:
                QMessageBox.warning(self, "Hiba", "Kérem adjon meg egy számot!")
                return
        
            value = float(text)
        
            if value < 0:
                QMessageBox.warning(self, "Hiba", "Kérem pozitív számot adjon meg!")
                return
            
            # Az aktuális zóna azonosítása
            current_zone = self.km_combo.currentText()
            current_zone_col = self.getZoneColumn(current_zone)
        
            # Az aktuális értékek tárolása
            if not hasattr(self, 'stored_values'):
                self.stored_values = {}
            
            # Az aktuális dátumhoz tartozó értékek frissítése
            current_date = self.date_edit.date().toString('yyyy-MM-dd')
            if current_date not in self.stored_values:
                self.stored_values[current_date] = {}
            
            # Az új érték hozzáadása a megfelelő övezethez tartozó listához
            if current_zone not in self.stored_values[current_date]:
                self.stored_values[current_date][current_zone] = []
        
            self.stored_values[current_date][current_zone].append(value)
        
            # Táblázat frissítése
            self.updateDeliveryTableWithStoredValues()
        
            # M3 input mező törlése
            self.m3_input.clear()
        
            # Összeg kijelző frissítése az aktuális értékekkel
            self.updateM3Sum(current_date, current_zone)
        
        except ValueError:
            QMessageBox.warning(self, "Hiba", "Kérem számot adjon meg (pl.: 6.0 vagy 6,0)")

    def updateM3Sum(self, current_date, current_zone):
        if current_date in self.stored_values and current_zone in self.stored_values[current_date]:
            values = self.stored_values[current_date][current_zone]
            values_text = " + ".join(f"{v:.1f}" for v in values)
            total = sum(values)
            self.m3_sum.setText(f"({values_text}) = {total:.1f}")
        else:
            self.m3_sum.setText("(0)")

    def updateDeliveryTableWithStoredValues(self):
        current_date = self.date_edit.date().toString('yyyy-MM-dd')
    
        if current_date in self.stored_values:
            # Megkeressük a megfelelő sort
            for row in range(self.delivery_table.rowCount()):
                if self.delivery_table.item(row, 0) and self.delivery_table.item(row, 0).text() == current_date:
                    # Minden tárolt érték frissítése a táblázatban
                    for zone, values in self.stored_values[current_date].items():
                        col = self.getZoneColumn(zone)
                        if col > 0:
                            # Az értékek megjelenítése + jellel elválasztva
                            display_text = " + ".join(f"{v:.1f}" for v in values)
                            new_item = QTableWidgetItem(display_text)
                            new_item.setTextAlignment(Qt.AlignCenter)
                            self.delivery_table.setItem(row, col, new_item)

    def getZoneColumn(self, zone_text):
        try:
            start_km = int(zone_text.split(' ')[1].split('-')[0])
            column = (start_km // 5) + 1
            return column
        except:
            return 0

    def saveWorkHours(self):
        try:
            data = {
                'date': self.date_edit.date().toString('yyyy-MM-dd'),
                'start_time': self.start_time.time().toString('HH:mm'),
                'end_time': self.end_time.time().toString('HH:mm'),
                'type': self.type_combo.currentText()
            }

            # Find the row for this date
            date_text = data['date']
            row_index = -1
            for row in range(self.work_table.rowCount()):
                if self.work_table.item(row, 0) and self.work_table.item(row, 0).text() == date_text:
                    row_index = row
                    break

            if row_index >= 0:
                work_type = data['type']
            
                # Clear previous entries in the row
                for col in range(2, 10):  # Clear all data columns except date and day
                    self.work_table.setItem(row_index, col, QTableWidgetItem(""))
            
                if work_type == "Sima munkanap":
                    # Sima munkanap - oszlopok 2,3,4
                    start_item = QTableWidgetItem(data['start_time'])
                    end_item = QTableWidgetItem(data['end_time'])
                    start_item.setTextAlignment(Qt.AlignCenter)
                    end_item.setTextAlignment(Qt.AlignCenter)
                    self.work_table.setItem(row_index, 2, start_item)  # Sima Kezdés
                    self.work_table.setItem(row_index, 3, end_item)    # Sima Végzés
                
                    # Órák számítása
                    start = QTime.fromString(data['start_time'], 'HH:mm')
                    end = QTime.fromString(data['end_time'], 'HH:mm')
                    if start.isValid() and end.isValid():
                        minutes = start.secsTo(end) / 60
                        hours = minutes / 60
                        hours_item = QTableWidgetItem(f"{hours:.2f}")
                        hours_item.setTextAlignment(Qt.AlignCenter)
                        self.work_table.setItem(row_index, 4, hours_item)  # Sima Ledolgozott

                elif work_type == "Műhely nap":
                    # Műhely nap - oszlopok 5,6,7
                    start_item = QTableWidgetItem(data['start_time'])
                    end_item = QTableWidgetItem(data['end_time'])
                    start_item.setTextAlignment(Qt.AlignCenter)
                    end_item.setTextAlignment(Qt.AlignCenter)
                    self.work_table.setItem(row_index, 5, start_item)  # Műhely Kezdés
                    self.work_table.setItem(row_index, 6, end_item)    # Műhely Végzés
                
                    # Órák számítása
                    start = QTime.fromString(data['start_time'], 'HH:mm')
                    end = QTime.fromString(data['end_time'], 'HH:mm')
                    if start.isValid() and end.isValid():
                        minutes = start.secsTo(end) / 60
                        hours = minutes / 60
                        hours_item = QTableWidgetItem(f"{hours:.2f}")
                        hours_item.setTextAlignment(Qt.AlignCenter)
                        self.work_table.setItem(row_index, 7, hours_item)  # Műhely Ledolgozott

                elif work_type == "Szabadság":
                    # Szabadság - oszlop 8
                    szabi_item = QTableWidgetItem("1")
                    szabi_item.setTextAlignment(Qt.AlignCenter)
                    self.work_table.setItem(row_index, 8, szabi_item)

                elif work_type == "Betegszabadság (TP)":
                    # Betegszabadság - oszlop 9
                    bszabi_item = QTableWidgetItem("1")
                    bszabi_item.setTextAlignment(Qt.AlignCenter)
                    self.work_table.setItem(row_index, 9, bszabi_item)

            # Save to JSON file
            with open('work_hours.json', 'a', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
                f.write('\n')

            # Update vacation days if needed
            if data['type'] == "Szabadság":
                self.updateVacationDays()

            QMessageBox.information(self, "Siker", "Munkaórák mentve!")

        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Hiba történt: {str(e)}")

    def saveDeliveryData(self):
        try:
            current_date = self.date_edit.date().toString('yyyy-MM-dd')
        
            # Ha nincsenek tárolt értékek, nincs mit menteni
            if not hasattr(self, 'stored_values') or current_date not in self.stored_values:
                QMessageBox.warning(self, "Figyelmeztetés", "Nincs mentendő adat!")
                return
            
            data = {
                'date': current_date,
                'factory': self.factory_combo.currentText(),
                'address': self.address_input.text(),
                'delivery_number': self.delivery_input.text(),
                'stored_values': self.stored_values[current_date]
            }
        
            # JSON fájl mentése
            with open('delivery_data.json', 'a', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
                f.write('\n')
        
            # Értékek megtartása a táblázatban
            self.updateDeliveryTableWithStoredValues()
        
            # M3 input és összeg kijelző törlése
            self.m3_input.clear()
            self.m3_sum.setText("(0)")
        
            # Tárolt értékek törlése a következő bevitelhez
            if current_date in self.stored_values:
                del self.stored_values[current_date]
        
            QMessageBox.information(self, "Siker", "Fuvar adatok mentve!")
        
        except Exception as e:
            print(f"Hiba részletek: {str(e)}")  # Debug információ
            QMessageBox.warning(self, "Hiba", f"Hiba történt: {str(e)}")

    def updateVacationDays(self):
        try:
            current_year = QDate.currentDate().year()
            cursor = self.conn.cursor()
        
            # Lekérdezzük az aktuális szabadság keretet
            cursor.execute("""
                SELECT total_days, used_days 
                FROM vacation_allowance 
                WHERE year = ?
            """, (current_year,))
        
            result = cursor.fetchone()
            if result:
                total_days, used_days = result
                # Növeljük a felhasznált napok számát
                used_days += 1
            
                # Frissítjük az adatbázist
                cursor.execute("""
                    UPDATE vacation_allowance 
                    SET used_days = ? 
                    WHERE year = ?
                """, (used_days, current_year))
            
                self.conn.commit()
            
                # Frissítjük a kijelzőt
                self.vacation_label.setText(f"Szabadság: {used_days}/{total_days}")
        
        except Exception as e:
            print(f"Hiba a szabadság frissítésekor: {str(e)}")
            QMessageBox.warning(self, "Hiba", f"Hiba a szabadság frissítésekor: {str(e)}")

    def exportToExcel(self):
        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Munkaórák"
            
            # Fejlécek
            headers = ["Dátum", "Nap", "Kezdés", "Végzés", "Munka típusa"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Adatok mentése
            for row in range(self.work_table.rowCount()):
                for col in range(len(headers)):
                    item = self.work_table.item(row, col)
                    if item:
                        ws1.cell(row=row+2, column=col+1, value=item.text())
            
            wb.save('munka_nyilvantartas.xlsx')
            QMessageBox.information(self, "Siker", "Excel fájl mentve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Mentési hiba: {str(e)}")

    def saveToExcel(self):
        try:
            print("Excel mentés kezdése...")
            wb = Workbook()
        
            # 1. Munkalap: Munkaórák
            print("Munkaórák munkalap létrehozása...")
            ws1 = wb.active
            ws1.title = "Munkaórák"
        
            # Munkaórák fejlécek
            work_headers = ["Dátum", "Nap", "Kezdés", "Végzés", "Munka típusa"]
            for col, header in enumerate(work_headers, 1):
                ws1.cell(row=1, column=col, value=header)
        
            # Munkaórák adatok
            print(f"Munkaóra sorok száma: {self.work_table.rowCount()}")
            for row in range(self.work_table.rowCount()):
                for col in range(len(work_headers)):
                    item = self.work_table.item(row, col)
                    if item:
                        ws1.cell(row=row+2, column=col+1, value=item.text())
        
            # 2. Munkalap: Fuvar adatok
            print("Fuvar adatok munkalap létrehozása...")
            ws2 = wb.create_sheet(title="Fuvar adatok")
        
            # Fuvar adatok fejlécek
            delivery_headers = ["Dátum"]
            for i in range(0, 45, 5):
                delivery_headers.append(f"Övezet {i}-{i+5}")
        
            print(f"Fuvar fejlécek: {delivery_headers}")
            for col, header in enumerate(delivery_headers, 1):
                ws2.cell(row=1, column=col, value=header)
        
            # Fuvar adatok
            print(f"Fuvar sorok száma: {self.delivery_table.rowCount()}")
            print(f"Fuvar oszlopok száma: {self.delivery_table.columnCount()}")
        
            for row in range(self.delivery_table.rowCount()):
                print(f"Fuvar sor {row} feldolgozása...")
                for col in range(self.delivery_table.columnCount()):
                    item = self.delivery_table.item(row, col)
                    if item:
                        value = item.text()
                        print(f"Cella érték: sor={row+2}, oszlop={col+1}, érték={value}")
                        ws2.cell(row=row+2, column=col+1, value=value)
        
            print("Fájl mentése...")
            wb.save('munka_nyilvantartas.xlsx')
            print("Fájl mentve!")
            QMessageBox.information(self, "Siker", "Excel fájl mentve!")
        
        except Exception as e:
            print(f"HIBA történt: {str(e)}")
            print(f"Hiba típusa: {type(e)}")
            QMessageBox.warning(self, "Hiba", f"Mentési hiba: {str(e)}")

    def saveWorkHoursToExcel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Munkaórák"

            # Fejlécek a képen látható módon
            headers = [
                "Dátum", "Nap", 
                "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
                "Műhely\nKezdés", "Műhely\nVégzés", "Ledolgozott\nÓrák",
                "Szabadság", "Betegszabadság\n(TP)"
            ]

            # Fejléc formázás
            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Fejlécek beállítása
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            # Adatok formázása
            data_font = Font(color="000000")
            data_alignment = Alignment(horizontal="center", vertical="center")
            data_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Adatok beillesztése és formázása
            for row in range(self.work_table.rowCount()):
                # Dátum és nap beállítása
                for col in [0, 1]:  # Dátum és nap oszlopok
                    cell = ws.cell(row=row + 2, column=col + 1)
                    item = self.work_table.item(row, col)
                    if item:
                        cell.value = item.text()
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border

                # Munkanap típus ellenőrzése és megfelelő cellák kitöltése
                work_type = None
                start_time = None
                end_time = None
                hours = None
            
                # Ellenőrizzük a munka típusát és időket
                for col in range(2, self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item and item.text():
                        if "Szabadság" in item.text():
                            ws.cell(row=row + 2, column=9).value = 1  # Szabadság oszlop
                        elif "Betegszabadság" in item.text():
                            ws.cell(row=row + 2, column=10).value = 1  # Betegszabadság oszlop
                        elif "Sima munkanap" in item.text():
                            start_time = self.work_table.item(row, 2).text() if self.work_table.item(row, 2) else None
                            end_time = self.work_table.item(row, 3).text() if self.work_table.item(row, 3) else None
                            hours = self.work_table.item(row, 4).text() if self.work_table.item(row, 4) else None
                        
                            if start_time:
                                ws.cell(row=row + 2, column=3).value = start_time
                            if end_time:
                                ws.cell(row=row + 2, column=4).value = end_time
                            if hours:
                                ws.cell(row=row + 2, column=5).value = float(hours)
                        elif "Műhely" in item.text():
                            start_time = self.work_table.item(row, 6).text() if self.work_table.item(row, 6) else None
                            end_time = self.work_table.item(row, 7).text() if self.work_table.item(row, 7) else None
                            hours = self.work_table.item(row, 8).text() if self.work_table.item(row, 8) else None
                        
                            if start_time:
                                ws.cell(row=row + 2, column=6).value = start_time
                            if end_time:
                                ws.cell(row=row + 2, column=7).value = end_time
                            if hours:
                                ws.cell(row=row + 2, column=8).value = float(hours)

                # Minden cella formázása az adott sorban
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row + 2, column=col)
                    if not cell.value:  # Ha üres a cella
                        cell.value = ""
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border

            # Oszlopszélességek beállítása
            column_widths = [12, 10, 15, 15, 12, 15, 15, 12, 10, 15]
            for i, width in enumerate(column_widths):
                ws.column_dimensions[chr(65 + i)].width = width

            # Weekend sorok kiemelése
            for row in range(2, ws.max_row + 1):
                day_name = ws.cell(row=row, column=2).value
                if day_name in ["Szombat", "Vasárnap"]:
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            wb.save('munkaora_nyilvantartas.xlsx')
            QMessageBox.information(self, "Siker", "Munkaórák mentve!")

        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Munkaóra mentési hiba: {str(e)}")

    def saveDeliveryToExcel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fuvar adatok"
        
            # Fejlécek
            headers = ["Dátum"]
            for i in range(0, 45, 5):
                headers.append(f"Övezet {i}-{i+5}")
        
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
            # Adatok
            for row in range(self.delivery_table.rowCount()):
                for col in range(len(headers)):
                    item = self.delivery_table.item(row, col)
                    if item:
                        ws.cell(row=row+2, column=col+1, value=item.text())
        
            # Oszlopszélességek beállítása
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        
            wb.save('fuvar_nyilvantartas.xlsx')
            QMessageBox.information(self, "Siker", "Fuvar adatok mentve!")
        
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Fuvar adat mentési hiba: {str(e)}")

    # Új kombinált mentés függvények
    def saveWorkHoursAndExport(self):
        self.saveWorkHours()
        self.saveWorkHoursToExcel()

    def saveDeliveryAndExport(self):
        self.saveDeliveryData()
        self.saveDeliveryToExcel()

    def openDatabaseManager(self):
        try:
            dbManager = DatabaseManager(self)
            dbManager.exec()
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba a törzsadat kezelő megnyitásakor: {str(e)}")

    def openExcel(self):
        try:
            file_name, _ = QFileDialog.getOpenFileName(
                self, "Excel fájl megnyitása", "", "Excel files (*.xlsx *.xls)"
            )
            if file_name:
                wb = load_workbook(file_name)
                self.loadDataFromExcel(wb)
                os.startfile(file_name)
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Fájl megnyitási hiba: {str(e)}")

    def loadDataFromExcel(self, workbook):
        try:
            # Munkaórák betöltése
            if "Munkaórák" in workbook.sheetnames:
                ws = workbook["Munkaórák"]
                for row in range(2, ws.max_row + 1):
                    date = ws.cell(row=row, column=1).value
                    if date:
                        self.loadWorkHoursRow(date, ws, row)
    
            # Fuvar adatok betöltése
            if "Fuvar adatok" in workbook.sheetnames:
                ws = workbook["Fuvar adatok"]
                for row in range(2, ws.max_row + 1):
                    date = ws.cell(row=row, column=1).value
                    if date:
                        self.loadDeliveryRow(date, ws, row)
                
            QMessageBox.information(self, "Siker", "Excel adatok betöltve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Adatok betöltési hiba: {str(e)}")

    def loadWorkHoursRow(self, date, ws, row):
        start_time = ws.cell(row=row, column=2).value
        end_time = ws.cell(row=row, column=3).value
        work_type = ws.cell(row=row, column=4).value
    
        for table_row in range(self.work_table.rowCount()):
            if self.work_table.item(table_row, 0).text() == str(date):
                if start_time:
                    self.work_table.setItem(table_row, 2, QTableWidgetItem(str(start_time)))
                if end_time:
                    self.work_table.setItem(table_row, 3, QTableWidgetItem(str(end_time)))
                if work_type:
                    self.work_table.setItem(table_row, 4, QTableWidgetItem(str(work_type)))

    def loadDeliveryRow(self, date, ws, row):
        for col in range(1, self.delivery_table.columnCount()):
            value = ws.cell(row=row, column=col+1).value
            if value is not None:
                for table_row in range(self.delivery_table.rowCount()):
                    if self.delivery_table.item(table_row, 0).text() == str(date):
                        self.delivery_table.setItem(table_row, col, QTableWidgetItem(str(value)))

    def printData(self):
        dialog = QPrintDialog(self)
        if dialog.exec() == QDialog.Accepted:
            printer = dialog.printer()
            # TODO: implement printing logic

    def updateVacationDisplay(self):
        try:
            cursor = self.conn.cursor()
            current_year = QDate.currentDate().year()
        
            cursor.execute("""
                SELECT total_days, used_days 
                FROM vacation_allowance 
                WHERE year = ?
            """, (current_year,))
        
            result = cursor.fetchone()
            if result:
                total_days, used_days = result
                self.vacation_label.setText(f"Szabadság: {used_days}/{total_days}")
            else:
                self.vacation_label.setText("Szabadság: 0/0")
            
        except Exception as e:
            print(f"Hiba a szabadság megjelenítésekor: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FuvarAdminApp()
    window.show()
    sys.exit(app.exec())  # Eltávolítottuk az aláhúzást